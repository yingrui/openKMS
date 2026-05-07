"""Build in-app preview JSON and markdown for .xlsx uploads (no VLM pipeline)."""

from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import load_workbook

DEFAULT_MAX_ROWS = 500
DEFAULT_MAX_COLS = 64


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    s = str(value).strip()
    return s


def _md_escape_cell(s: str) -> str:
    return re.sub(r"[\r\n]+", " ", s).replace("|", "\\|")


def _trim_trailing_empty(cells: list[str]) -> list[str]:
    out = list(cells)
    while out and out[-1] == "":
        out.pop()
    return out


def build_xlsx_preview(content: bytes, *, file_hash: str, max_rows: int = DEFAULT_MAX_ROWS, max_cols: int = DEFAULT_MAX_COLS) -> tuple[dict[str, Any], str]:
    """Return ``(parsing_result, markdown)`` for storage on the document row."""
    bio = io.BytesIO(content)
    wb = load_workbook(filename=bio, read_only=True, data_only=True)
    try:
        sheets_out: list[dict[str, Any]] = []
        md_parts: list[str] = []
        for ws in wb.worksheets:
            rows_out: list[list[str]] = []
            truncated_rows = False
            try:
                truncated_cols = bool(ws.max_column and ws.max_column > max_cols)
            except Exception:
                truncated_cols = False
            for i, row in enumerate(ws.iter_rows(min_row=1, max_col=max_cols, values_only=True)):
                if i >= max_rows:
                    truncated_rows = True
                    break
                rows_out.append(_trim_trailing_empty([_cell_str(c) for c in row]))
            sheets_out.append(
                {
                    "name": ws.title,
                    "rows": rows_out,
                    "truncated_rows": truncated_rows,
                    "truncated_cols": truncated_cols,
                    "max_rows": max_rows,
                    "max_cols": max_cols,
                }
            )
            md_parts.append(f"## {ws.title}\n")
            if not rows_out:
                md_parts.append("(empty)\n\n")
                continue
            header = rows_out[0]
            md_parts.append("| " + " | ".join(_md_escape_cell(c) for c in header) + " |\n")
            md_parts.append("| " + " | ".join("---" for _ in header) + " |\n")
            for data_row in rows_out[1:]:
                padded = list(data_row) + [""] * (len(header) - len(data_row))
                trimmed = padded[: len(header)]
                md_parts.append("| " + " | ".join(_md_escape_cell(c) for c in trimmed) + " |\n")
            md_parts.append("\n")
        preview: dict[str, Any] = {
            "document_kind": "spreadsheet",
            "file_hash": file_hash,
            "page_count": len(sheets_out),
            "sheets": sheets_out,
        }
        return preview, "".join(md_parts).strip()
    finally:
        wb.close()
